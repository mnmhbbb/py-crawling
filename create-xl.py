import openpyxl

# 1. 엑셀 만들기
wb = openpyxl.Workbook()

# 2. 엑셀 워크시트 만들기
ws = wb.create_sheet('worksheet-1')

# 3. 데이터 추가하기
ws['A2'] = '번호'
ws['B2'] = '이름'

ws['A3'] = '123'
ws['B3'] = '테스트'

# 4. 엑셀 저장하기
wb.save('test-1.xlsx') # 현재 파일 경로에 생성
wb.save(r'/Users/mhbaek/dev/python/xlsx/test-1.xlsx') 